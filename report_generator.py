"""
report_generator.py — Generates HTML report (Chart.js), PDF (matplotlib + Playwright),
and Excel (openpyxl) deliverables for BenchmarkHub.
"""

import os
import json
import asyncio
import logging
from datetime import datetime
from typing import Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import base64
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Design system colors
COLORS = {
    "primary": "#1B2A4A",
    "secondary": "#2E86AB",
    "accent": "#F18F01",
    "success": "#2BA84A",
    "danger": "#E63946",
    "bg": "#F8F9FA",
    "card_bg": "#FFFFFF",
    "text": "#2D3436",
}

LOGO_URL = "https://www.educaedtech.com/images/educa-edtech-group-dark.webp"


def fig_to_base64(fig, dpi=150):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight", facecolor="white")
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode("utf-8")
    plt.close(fig)
    return f"data:image/png;base64,{b64}"


# ─────────────── HTML REPORT (Chart.js) ───────────────

def generate_html_report(
    analysis_data: dict,
    research_results: list,
    strategic_data: dict,
    proposals: dict,
    faculty_name: str = "General",
    output_path: str = "informe_benchmark.html",
) -> str:
    """Generate interactive HTML report with Chart.js."""

    kpis = analysis_data.get("kpis", {})
    top20 = analysis_data.get("top_20", [])
    emerging = analysis_data.get("emerging", [])
    declining = analysis_data.get("declining", [])
    by_faculty = analysis_data.get("by_faculty", [])
    year_cols = analysis_data.get("year_columns", [])
    name_col = analysis_data.get("name_column", "Producto")
    swot = strategic_data.get("swot", {})
    competitor_stars = strategic_data.get("competitor_stars", [])
    improvements = proposals.get("improvements", [])
    new_products = proposals.get("new_products", [])

    # Prepare chart data
    top20_names = [_truncate(r.get(name_col, "?"), 40) for r in top20[:15]]
    top20_values = [r.get("ventas_total", 0) for r in top20[:15]]

    # Faculty donut
    fac_labels = [r.get("Facultad", "?") for r in by_faculty[:8]]
    fac_values = [r.get("ventas_totales", 0) for r in by_faculty[:8]]
    if len(by_faculty) > 8:
        fac_labels.append("Otros")
        fac_values.append(sum(r.get("ventas_totales", 0) for r in by_faculty[8:]))

    # Growth data
    growth_names = []
    growth_values = []
    for r in (emerging[:5] + declining[:5]):
        growth_names.append(_truncate(r.get(name_col, "?"), 30))
        growth_values.append(r.get("crecimiento_pct", 0))

    # Trend data
    trend_data = {}
    for r in top20[:5]:
        pname = _truncate(r.get(name_col, "?"), 25)
        trend_data[pname] = [r.get(yc, 0) for yc in year_cols]

    # Competition comparison
    comp_our_prices = []
    comp_avg_prices = []
    comp_labels = []
    for res in research_results:
        if res.get("competitors"):
            pname = _truncate(res.get("our_product", "?"), 25)
            comp_labels.append(pname)
            # Find our price from products
            our_price = 0
            for t in top20:
                if t.get(name_col, "") == res.get("our_product", ""):
                    our_price = t.get("Precio", 0) or 0
                    break
            comp_our_prices.append(our_price)
            prices = []
            for c in res["competitors"]:
                try:
                    p = float(str(c.get("price", "0")).replace("€", "").replace(",", ".").strip())
                    if p > 0:
                        prices.append(p)
                except (ValueError, TypeError):
                    pass
            comp_avg_prices.append(round(sum(prices) / len(prices)) if prices else 0)

    # Competition table
    comp_rows_html = ""
    for res in research_results:
        for comp in res.get("competitors", []):
            url = comp.get("url", "#")
            comp_rows_html += f"""<tr>
                <td>{res.get('our_product', '')}</td>
                <td>{comp.get('competitor_name', '')}</td>
                <td>{comp.get('product_name', '')}</td>
                <td>{comp.get('price', 'N/D')}</td>
                <td>{comp.get('hours', 'N/D')}</td>
                <td>{comp.get('ects', 'N/D')}</td>
                <td>{comp.get('degree_type', 'N/D')}</td>
                <td>{comp.get('value_attributes', '')}</td>
                <td>{comp.get('key_differentiator', '')}</td>
                <td><a href="{url}" target="_blank" rel="noopener">Ver</a></td>
            </tr>"""

    # SWOT HTML
    def _swot_list(items):
        return "".join(f"<li>{i}</li>" for i in (items or []))

    # Competitor stars HTML
    stars_html = ""
    for s in competitor_stars:
        cls = s.get("classification", "")
        emoji = {"amenaza_directa": "&#128308;", "oportunidad_nicho": "&#128993;", "referente_calidad": "&#128994;"}.get(cls, "")
        label = {"amenaza_directa": "Amenaza directa", "oportunidad_nicho": "Oportunidad", "referente_calidad": "Referente"}.get(cls, cls)
        stars_html += f"""<tr>
            <td>{emoji} {label}</td>
            <td>{s.get('competitor', '')}</td>
            <td>{s.get('product', '')}</td>
            <td>{s.get('reason', '')}</td>
            <td>{s.get('impact', '')}</td>
        </tr>"""

    # Proposals HTML
    improvements_html = ""
    for imp in improvements:
        prio = imp.get("priority", "media")
        prio_color = {"alta": "#E63946", "media": "#F18F01", "baja": "#2BA84A"}.get(prio, "#999")
        prio_label = {"alta": "&#128308; Alta", "media": "&#128993; Media", "baja": "&#128994; Baja"}.get(prio, prio)
        improvements_html += f"""<tr>
            <td style="color:{prio_color};font-weight:bold">{prio_label}</td>
            <td>{imp.get('current_product', '')}</td>
            <td>{imp.get('current_price', '')} &rarr; {imp.get('proposed_price', '')}</td>
            <td>{imp.get('proposed_name', '-')}</td>
            <td>{imp.get('proposed_hours_ects', '-')}</td>
            <td>{', '.join(imp.get('attributes_to_add', []))}</td>
            <td>{imp.get('strategic_justification', '')}</td>
        </tr>"""

    new_products_html = ""
    for i, np in enumerate(new_products, 1):
        prio = np.get("priority", "media")
        prio_color = {"alta": "#E63946", "media": "#F18F01", "baja": "#2BA84A"}.get(prio, "#999")
        prio_label = {"alta": "&#128308; Alta", "media": "&#128993; Media", "baja": "&#128994; Baja"}.get(prio, prio)
        new_products_html += f"""<tr>
            <td>{i}</td>
            <td style="color:{prio_color};font-weight:bold">{prio_label}</td>
            <td>{np.get('name', '')}</td>
            <td>{np.get('type', '')}</td>
            <td>{np.get('recommended_price', '')}</td>
            <td>{np.get('hours_ects', '')}</td>
            <td>{', '.join(np.get('key_attributes', []))}</td>
            <td>{np.get('strategic_justification', '')}</td>
        </tr>"""

    # Year KPI cards
    year_kpi_html = ""
    for yc in year_cols[-4:]:
        val = kpis.get(f"ventas_{yc}", 0)
        year_kpi_html += f"""
        <div class="kpi-card">
            <div class="kpi-icon">&#128197;</div>
            <div class="kpi-value">{val:,}</div>
            <div class="kpi-label">Ventas {yc}</div>
        </div>"""

    today = datetime.now().strftime("%d/%m/%Y")

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Benchmark de Producto - {faculty_name} | EDUCA EDTECH Group</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: 'Segoe UI', system-ui, -apple-system, sans-serif; background: {COLORS['bg']}; color: {COLORS['text']}; line-height: 1.6; }}
header {{ background: linear-gradient(135deg, {COLORS['primary']}, {COLORS['secondary']}); color: white; padding: 30px 40px; }}
header img {{ height: 40px; margin-bottom: 12px; }}
header h1 {{ font-size: 1.8rem; font-weight: 700; }}
header p {{ opacity: 0.9; font-size: 0.95rem; }}
.container {{ max-width: 1400px; margin: 0 auto; padding: 20px 30px; }}
.section {{ background: {COLORS['card_bg']}; border-radius: 12px; padding: 24px; margin-bottom: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
.section h2 {{ color: {COLORS['primary']}; font-size: 1.3rem; margin-bottom: 16px; border-bottom: 3px solid {COLORS['secondary']}; padding-bottom: 8px; }}
.kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 16px; margin-bottom: 20px; }}
.kpi-card {{ background: white; border: 1px solid #e0e0e0; border-radius: 10px; padding: 20px; text-align: center; }}
.kpi-icon {{ font-size: 1.5rem; margin-bottom: 4px; }}
.kpi-value {{ font-size: 1.8rem; font-weight: 700; color: {COLORS['primary']}; }}
.kpi-label {{ font-size: 0.85rem; color: #666; margin-top: 4px; }}
.chart-container {{ position: relative; margin: 16px 0; }}
.chart-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
@media (max-width: 900px) {{ .chart-row {{ grid-template-columns: 1fr; }} }}
table {{ width: 100%; border-collapse: collapse; font-size: 0.85rem; }}
th {{ background: {COLORS['primary']}; color: white; padding: 10px 8px; text-align: left; position: sticky; top: 0; }}
td {{ padding: 8px; border-bottom: 1px solid #eee; }}
tr:hover {{ background: #f0f7ff; }}
.table-scroll {{ overflow-x: auto; }}
.table-scroll table {{ min-width: 1200px; }}
a {{ color: {COLORS['secondary']}; }}
/* SWOT grid */
.swot-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 2px; background: #ddd; border-radius: 8px; overflow: hidden; }}
.swot-box {{ padding: 20px; }}
.swot-box h3 {{ margin-bottom: 10px; font-size: 1rem; }}
.swot-box ul {{ padding-left: 20px; }}
.swot-box li {{ margin-bottom: 6px; font-size: 0.88rem; }}
.swot-s {{ background: #d4edda; }}
.swot-w {{ background: #fff3cd; }}
.swot-o {{ background: #d6eaf8; }}
.swot-t {{ background: #f8d7da; }}
.priority-tag {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-weight: 600; font-size: 0.8rem; }}
footer {{ background: {COLORS['primary']}; color: white; padding: 20px 40px; margin-top: 30px; font-size: 0.8rem; opacity: 0.9; }}
@media print {{
  body {{ background: white; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
  .kpi-card, .chart-section, .section {{ break-inside: avoid; }}
  header {{ background: {COLORS['primary']} !important; }}
  canvas {{ max-width: 100% !important; height: auto !important; }}
}}
</style>
</head>
<body>
<header>
    <img src="{LOGO_URL}" alt="EDUCA EDTECH Group" onerror="this.style.display='none'">
    <h1>Informe de Benchmark de Producto Formativo</h1>
    <p>Facultad: {faculty_name} &nbsp;|&nbsp; Fecha: {today} &nbsp;|&nbsp; EDUCA EDTECH Group</p>
</header>

<div class="container">

<!-- KPIs -->
<div class="section">
    <h2>Resumen Ejecutivo</h2>
    <div class="kpi-grid">
        <div class="kpi-card">
            <div class="kpi-icon">&#128200;</div>
            <div class="kpi-value">{kpis.get('total_ventas', 0):,}</div>
            <div class="kpi-label">Ventas Totales</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-icon">&#128218;</div>
            <div class="kpi-value">{kpis.get('total_productos', 0)}</div>
            <div class="kpi-label">Total Productos</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-icon">&#9989;</div>
            <div class="kpi-value">{kpis.get('productos_activos', 0)}</div>
            <div class="kpi-label">Productos Activos</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-icon">&#128308;</div>
            <div class="kpi-value">{kpis.get('productos_muertos', 0)}</div>
            <div class="kpi-label">Productos Muertos</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-icon">&#127942;</div>
            <div class="kpi-value" style="font-size:1rem">{_truncate(kpis.get('producto_top', ''), 30)}</div>
            <div class="kpi-label">Producto Top ({kpis.get('ventas_producto_top', 0):,} ventas)</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-icon">&#128200;</div>
            <div class="kpi-value">{kpis.get('crecimiento_medio', 0)}%</div>
            <div class="kpi-label">Crecimiento Medio</div>
        </div>
        {year_kpi_html}
    </div>
</div>

<!-- Charts: Top 20 + Faculty -->
<div class="section">
    <h2>Analisis de Ventas</h2>
    <div class="chart-row">
        <div class="chart-container"><canvas id="chartTop20"></canvas></div>
        <div class="chart-container"><canvas id="chartFaculty"></canvas></div>
    </div>
    <div class="chart-row" style="margin-top:20px">
        <div class="chart-container"><canvas id="chartGrowth"></canvas></div>
        <div class="chart-container"><canvas id="chartTrend"></canvas></div>
    </div>
</div>

<!-- Competition Table -->
<div class="section">
    <h2>Mapa Competitivo</h2>
    {"<div class='chart-container'><canvas id='chartPriceComp'></canvas></div>" if comp_labels else ""}
    <div class="table-scroll">
        <table>
            <tr>
                <th>Nuestro Producto</th><th>Competidor</th><th>Producto Competidor</th>
                <th>Precio</th><th>Horas</th><th>ECTS</th><th>Titulacion</th>
                <th>Atributos de Valor</th><th>Diferenciador</th><th>URL</th>
            </tr>
            {comp_rows_html}
        </table>
    </div>
</div>

<!-- Competitor Stars -->
{"<div class='section'><h2>Productos Estrella de Competidores</h2><div class='table-scroll'><table><tr><th>Clasificacion</th><th>Competidor</th><th>Producto</th><th>Razon</th><th>Impacto</th></tr>" + stars_html + "</table></div></div>" if stars_html else ""}

<!-- SWOT -->
<div class="section">
    <h2>Analisis DAFO</h2>
    <div class="swot-grid">
        <div class="swot-box swot-s"><h3>&#128170; Fortalezas</h3><ul>{_swot_list(swot.get('strengths', []))}</ul></div>
        <div class="swot-box swot-w"><h3>&#9888;&#65039; Debilidades</h3><ul>{_swot_list(swot.get('weaknesses', []))}</ul></div>
        <div class="swot-box swot-o"><h3>&#128161; Oportunidades</h3><ul>{_swot_list(swot.get('opportunities', []))}</ul></div>
        <div class="swot-box swot-t"><h3>&#128680; Amenazas</h3><ul>{_swot_list(swot.get('threats', []))}</ul></div>
    </div>
</div>

<!-- Proposals: Improvements -->
<div class="section">
    <h2>Propuestas de Mejora</h2>
    <div class="table-scroll">
        <table>
            <tr><th>Prioridad</th><th>Producto</th><th>Precio Actual &rarr; Propuesto</th><th>Nombre Propuesto</th><th>Horas/ECTS</th><th>Atributos a Anadir</th><th>Justificacion</th></tr>
            {improvements_html}
        </table>
    </div>
</div>

<!-- Proposals: New Products -->
<div class="section">
    <h2>Nuevos Productos Propuestos</h2>
    <div class="table-scroll">
        <table>
            <tr><th>#</th><th>Prioridad</th><th>Producto</th><th>Tipo</th><th>Precio</th><th>Horas/ECTS</th><th>Atributos Clave</th><th>Justificacion</th></tr>
            {new_products_html}
        </table>
    </div>
</div>

</div><!-- /container -->

<footer>
    <p><strong>EDUCA EDTECH Group</strong> &mdash; Uso interno</p>
    <p>Informe generado el {today}. Los precios de competencia pueden cambiar; se recomienda actualizar cada 1-3 meses.</p>
    <p>Generado con BenchmarkHub.</p>
</footer>

<script>
{_build_chartjs_script(top20_names, top20_values, fac_labels, fac_values, growth_names, growth_values, trend_data, year_cols, comp_labels, comp_our_prices, comp_avg_prices)}
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return output_path


def _build_chartjs_script(top20_names, top20_values, fac_labels, fac_values,
                          growth_names, growth_values, trend_data, year_cols,
                          comp_labels, comp_our_prices, comp_avg_prices) -> str:
    """Build all Chart.js initialization code as a plain string (no nested f-strings)."""
    primary = COLORS["primary"]
    secondary = COLORS["secondary"]
    accent = COLORS["accent"]
    success = COLORS["success"]
    danger = COLORS["danger"]

    parts = []

    # Top 20 bar chart
    parts.append(
        "new Chart(document.getElementById('chartTop20'), {"
        "type: 'bar',"
        "data: {labels: " + json.dumps(top20_names) + ","
        "datasets: [{label: 'Ventas Totales', data: " + json.dumps(top20_values) + ", backgroundColor: '" + secondary + "'}]},"
        "options: {indexAxis: 'y', responsive: true,"
        "plugins: {title: {display: true, text: 'Top 15 Productos por Ventas', color: '" + primary + "', font: {size: 14, weight: 'bold'}}, legend: {display: false}},"
        "scales: {x: {beginAtZero: true}}}});"
    )

    # Faculty donut
    if fac_labels:
        parts.append(
            "new Chart(document.getElementById('chartFaculty'), {"
            "type: 'doughnut',"
            "data: {labels: " + json.dumps(fac_labels) + ","
            "datasets: [{data: " + json.dumps(fac_values) + ", backgroundColor: ['#2E86AB','#F18F01','#2BA84A','#E63946','#9B59B6','#1ABC9C','#E67E22','#3498DB','#95A5A6']}]},"
            "options: {responsive: true,"
            "plugins: {title: {display: true, text: 'Distribucion por Facultad', color: '" + primary + "', font: {size: 14, weight: 'bold'}}}}});"
        )

    # Growth chart
    if growth_names:
        parts.append(
            "new Chart(document.getElementById('chartGrowth'), {"
            "type: 'bar',"
            "data: {labels: " + json.dumps(growth_names) + ","
            "datasets: [{label: 'Crecimiento %', data: " + json.dumps(growth_values) + ","
            "backgroundColor: " + json.dumps(growth_values) + ".map(v => v > 0 ? '" + success + "' : '" + danger + "')}]},"
            "options: {responsive: true,"
            "plugins: {title: {display: true, text: 'Crecimiento Interanual (%)', color: '" + primary + "', font: {size: 14, weight: 'bold'}}, legend: {display: false}}}});"
        )

    # Trend line
    parts.append(_generate_trend_chart_js(trend_data, year_cols))

    # Price comparison
    if comp_labels:
        parts.append(
            "new Chart(document.getElementById('chartPriceComp'), {"
            "type: 'bar',"
            "data: {labels: " + json.dumps(comp_labels) + ","
            "datasets: [{label: 'Nuestro Precio', data: " + json.dumps(comp_our_prices) + ", backgroundColor: '" + secondary + "'},"
            "{label: 'Media Competencia', data: " + json.dumps(comp_avg_prices) + ", backgroundColor: '" + accent + "'}]},"
            "options: {responsive: true,"
            "plugins: {title: {display: true, text: 'Comparativa de Precios vs Competencia', color: '" + primary + "', font: {size: 14, weight: 'bold'}}}}});"
        )

    return "\n".join(parts)


def _generate_trend_chart_js(trend_data: dict, year_cols: list) -> str:
    if not trend_data or not year_cols:
        return ""
    colors = ["#2E86AB", "#F18F01", "#2BA84A", "#E63946", "#9B59B6"]
    datasets = []
    for i, (name, values) in enumerate(trend_data.items()):
        c = colors[i % len(colors)]
        datasets.append({
            "label": name,
            "data": values,
            "borderColor": c,
            "tension": 0.3,
            "fill": False,
        })
    return f"""new Chart(document.getElementById('chartTrend'), {{
    type: 'line',
    data: {{
        labels: {json.dumps(year_cols)},
        datasets: {json.dumps(datasets)}
    }},
    options: {{
        responsive: true,
        plugins: {{ title: {{ display: true, text: 'Tendencia Top 5 Productos', color: '{COLORS["primary"]}', font: {{ size: 14, weight: 'bold' }} }} }}
    }}
}});"""


# ─────────────── PDF REPORT (matplotlib) ───────────────

async def generate_pdf_report(
    analysis_data: dict,
    research_results: list,
    strategic_data: dict,
    proposals: dict,
    faculty_name: str = "General",
    output_path: str = "informe_benchmark.pdf",
) -> str:
    """Generate PDF via matplotlib charts + Playwright."""

    kpis = analysis_data.get("kpis", {})
    top20 = analysis_data.get("top_20", [])
    year_cols = analysis_data.get("year_columns", [])
    name_col = analysis_data.get("name_column", "Producto")
    by_faculty = analysis_data.get("by_faculty", [])
    emerging = analysis_data.get("emerging", [])
    declining = analysis_data.get("declining", [])
    swot = strategic_data.get("swot", {})
    improvements = proposals.get("improvements", [])
    new_products = proposals.get("new_products", [])

    charts_b64 = {}

    # Chart 1: Top 15 horizontal bar
    fig, ax = plt.subplots(figsize=(10, 5.5))
    names = [_truncate(r.get(name_col, "?"), 35) for r in top20[:15]][::-1]
    vals = [r.get("ventas_total", 0) for r in top20[:15]][::-1]
    ax.barh(names, vals, color="#2E86AB")
    ax.set_title("Top 15 Productos por Ventas", color="#1B2A4A", weight="bold", fontsize=13)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x):,}"))
    plt.tight_layout()
    charts_b64["top15"] = fig_to_base64(fig)

    # Chart 2: Faculty pie
    if by_faculty:
        fig, ax = plt.subplots(figsize=(6, 4))
        labels = [r.get("Facultad", "?") for r in by_faculty[:8]]
        sizes = [r.get("ventas_totales", 0) for r in by_faculty[:8]]
        colors_pie = ["#2E86AB", "#F18F01", "#2BA84A", "#E63946", "#9B59B6", "#1ABC9C", "#E67E22", "#3498DB"]
        ax.pie(sizes, labels=labels, colors=colors_pie[:len(labels)], autopct="%1.0f%%", startangle=90, textprops={"fontsize": 8})
        ax.set_title("Distribucion por Facultad", color="#1B2A4A", weight="bold", fontsize=12)
        charts_b64["faculty"] = fig_to_base64(fig)

    # Chart 3: Growth
    growth_items = (emerging[:5] + declining[:5])
    if growth_items:
        fig, ax = plt.subplots(figsize=(10, 4))
        gnames = [_truncate(r.get(name_col, "?"), 30) for r in growth_items]
        gvals = [r.get("crecimiento_pct", 0) for r in growth_items]
        gcolors = ["#2BA84A" if v > 0 else "#E63946" for v in gvals]
        ax.bar(range(len(gnames)), gvals, color=gcolors)
        ax.set_xticks(range(len(gnames)))
        ax.set_xticklabels(gnames, rotation=45, ha="right", fontsize=7)
        ax.set_title("Crecimiento Interanual (%)", color="#1B2A4A", weight="bold", fontsize=12)
        ax.axhline(y=0, color="#999", linewidth=0.5)
        plt.tight_layout()
        charts_b64["growth"] = fig_to_base64(fig)

    today = datetime.now().strftime("%d/%m/%Y")

    # Build PDF HTML
    kpi_html = f"""
    <div class="kpi-row">
        <div class="kpi">Ventas Totales<br><strong>{kpis.get('total_ventas', 0):,}</strong></div>
        <div class="kpi">Productos<br><strong>{kpis.get('total_productos', 0)}</strong></div>
        <div class="kpi">Activos<br><strong>{kpis.get('productos_activos', 0)}</strong></div>
        <div class="kpi">Muertos<br><strong>{kpis.get('productos_muertos', 0)}</strong></div>
        <div class="kpi">Crecimiento Medio<br><strong>{kpis.get('crecimiento_medio', 0)}%</strong></div>
    </div>"""

    swot_html = ""
    for section, title, color in [
        ("strengths", "Fortalezas", "#d4edda"),
        ("weaknesses", "Debilidades", "#fff3cd"),
        ("opportunities", "Oportunidades", "#d6eaf8"),
        ("threats", "Amenazas", "#f8d7da"),
    ]:
        items = swot.get(section, [])
        items_html = "".join(f"<li>{i}</li>" for i in items)
        swot_html += f'<div style="background:{color};padding:15px;border-radius:6px"><h3>{title}</h3><ul>{items_html}</ul></div>'

    # Proposals tables
    imp_rows = ""
    for imp in improvements:
        imp_rows += f"<tr><td>{imp.get('priority','')}</td><td>{imp.get('current_product','')}</td><td>{imp.get('current_price','')} -> {imp.get('proposed_price','')}</td><td>{imp.get('proposed_name','-')}</td><td>{', '.join(imp.get('attributes_to_add',[]))}</td></tr>"

    new_rows = ""
    for np_ in new_products:
        new_rows += f"<tr><td>{np_.get('name','')}</td><td>{np_.get('type','')}</td><td>{np_.get('recommended_price','')}</td><td>{np_.get('hours_ects','')}</td><td>{np_.get('strategic_justification','')}</td></tr>"

    pdf_html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
body {{ font-family: 'DejaVu Sans', 'Segoe UI', sans-serif; margin: 20px; color: #2D3436; font-size: 11px; }}
h1 {{ color: white; background: #1B2A4A; padding: 20px; margin: -20px -20px 20px -20px; }}
h2 {{ color: #1B2A4A; border-bottom: 2px solid #2E86AB; padding-bottom: 5px; margin-top: 25px; }}
.kpi-row {{ display: flex; gap: 15px; margin: 15px 0; }}
.kpi {{ background: #f0f0f0; padding: 12px; border-radius: 6px; text-align: center; flex: 1; }}
.kpi strong {{ font-size: 1.4em; color: #1B2A4A; display: block; }}
img {{ max-width: 100%; height: auto; }}
table {{ width: 100%; border-collapse: collapse; margin: 10px 0; font-size: 10px; }}
th {{ background: #1B2A4A; color: white; padding: 6px; text-align: left; }}
td {{ padding: 5px; border-bottom: 1px solid #ddd; }}
.swot-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin: 15px 0; }}
.swot-grid h3 {{ margin: 0 0 8px 0; font-size: 12px; }}
.swot-grid ul {{ padding-left: 18px; margin: 0; }}
.swot-grid li {{ margin-bottom: 4px; }}
.page-break {{ page-break-before: always; }}
footer {{ margin-top: 30px; padding-top: 10px; border-top: 1px solid #ccc; font-size: 9px; color: #666; }}
</style></head><body>
<h1>Benchmark de Producto - {faculty_name}</h1>
<p>EDUCA EDTECH Group | {today}</p>

<h2>Resumen Ejecutivo</h2>
{kpi_html}

<h2>Top Productos</h2>
<img src="{charts_b64.get('top15', '')}">

{f'<div class="page-break"></div><h2>Distribucion por Facultad</h2><img src="{charts_b64["faculty"]}">' if "faculty" in charts_b64 else ""}

{f'<h2>Crecimiento Interanual</h2><img src="{charts_b64["growth"]}">' if "growth" in charts_b64 else ""}

<div class="page-break"></div>
<h2>Analisis DAFO</h2>
<div class="swot-grid">{swot_html}</div>

<div class="page-break"></div>
<h2>Propuestas de Mejora</h2>
<table><tr><th>Prioridad</th><th>Producto</th><th>Precio</th><th>Nombre Propuesto</th><th>Atributos</th></tr>{imp_rows}</table>

<h2>Nuevos Productos</h2>
<table><tr><th>Producto</th><th>Tipo</th><th>Precio</th><th>Horas/ECTS</th><th>Justificacion</th></tr>{new_rows}</table>

<footer>
<p>EDUCA EDTECH Group - Uso interno. Precios de competencia pueden cambiar; actualizar cada 1-3 meses.</p>
<p>Generado con BenchmarkHub el {today}.</p>
</footer>
</body></html>"""

    # Write temp HTML, convert to PDF
    html_path = output_path.replace(".pdf", "_temp.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(pdf_html)

    try:
        from playwright.async_api import async_playwright
        async with async_playwright() as p:
            browser = await p.chromium.launch(args=["--no-sandbox"])
            page = await browser.new_page(viewport={"width": 1200, "height": 900})
            await page.goto(f"file://{os.path.abspath(html_path)}", wait_until="load")
            await page.wait_for_timeout(1000)
            await page.pdf(
                path=output_path,
                format="A4",
                landscape=True,
                print_background=True,
                margin={"top": "10mm", "bottom": "10mm", "left": "8mm", "right": "8mm"},
            )
            await browser.close()
        # Clean up temp
        if os.path.exists(html_path):
            os.remove(html_path)
    except Exception as e:
        logger.error(f"PDF generation error: {e}")
        raise

    return output_path


# ─────────────── EXCEL REPORT (openpyxl) ───────────────

def generate_excel_report(
    analysis_data: dict,
    research_results: list,
    strategic_data: dict,
    proposals: dict,
    faculty_name: str = "General",
    output_path: str = "benchmark_datos.xlsx",
) -> str:
    """Generate professional Excel with 9 sheets."""

    wb = openpyxl.Workbook()
    kpis = analysis_data.get("kpis", {})
    top20 = analysis_data.get("top_20", [])
    year_cols = analysis_data.get("year_columns", [])
    name_col = analysis_data.get("name_column", "Producto")
    swot = strategic_data.get("swot", {})
    competitor_stars = strategic_data.get("competitor_stars", [])
    improvements = proposals.get("improvements", [])
    new_products = proposals.get("new_products", [])

    # Styles
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    normal_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    green_fill = PatternFill(start_color="D4EDDA", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", fill_type="solid")
    blue_fill = PatternFill(start_color="D6EAF8", fill_type="solid")
    orange_fill = PatternFill(start_color="FFF3CD", fill_type="solid")

    def _style_header(ws, cols):
        for i, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
        ws.freeze_panes = "A2"

    def _style_cell(cell):
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _auto_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    val = str(cell.value or "")
                    max_len = max(max_len, min(len(val), 50))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    # ── Sheet 1: Resumen Ejecutivo ──
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.cell(row=1, column=1, value="Benchmark de Producto Formativo").font = Font(name="Arial", size=14, bold=True, color="1B2A4A")
    ws1.cell(row=2, column=1, value=f"Facultad: {faculty_name}").font = Font(name="Arial", size=11)
    ws1.cell(row=3, column=1, value=f"Fecha: {datetime.now().strftime('%d/%m/%Y')}").font = normal_font

    kpi_items = [
        ("Ventas Totales", kpis.get("total_ventas", 0)),
        ("Total Productos", kpis.get("total_productos", 0)),
        ("Productos Activos", kpis.get("productos_activos", 0)),
        ("Productos Muertos", kpis.get("productos_muertos", 0)),
        ("Producto Top", kpis.get("producto_top", "")),
        ("Ventas Producto Top", kpis.get("ventas_producto_top", 0)),
        ("Crecimiento Medio", f"{kpis.get('crecimiento_medio', 0)}%"),
    ]
    for yc in year_cols:
        kpi_items.append((f"Ventas {yc}", kpis.get(f"ventas_{yc}", 0)))

    _style_header(ws1, ["KPI", "Valor"])
    ws1.delete_rows(1)
    # Re-add title rows
    ws1.insert_rows(1, 4)
    ws1.cell(row=1, column=1, value="Benchmark de Producto Formativo").font = Font(name="Arial", size=14, bold=True, color="1B2A4A")
    ws1.cell(row=2, column=1, value=f"Facultad: {faculty_name}").font = Font(name="Arial", size=11)
    ws1.cell(row=3, column=1, value=f"Fecha: {datetime.now().strftime('%d/%m/%Y')}").font = normal_font

    headers_row = 5
    ws1.cell(row=headers_row, column=1, value="KPI").font = header_font
    ws1.cell(row=headers_row, column=1).fill = header_fill
    ws1.cell(row=headers_row, column=2, value="Valor").font = header_font
    ws1.cell(row=headers_row, column=2).fill = header_fill

    for i, (k, v) in enumerate(kpi_items, headers_row + 1):
        c1 = ws1.cell(row=i, column=1, value=k)
        c2 = ws1.cell(row=i, column=2, value=v)
        _style_cell(c1)
        _style_cell(c2)
        c1.font = Font(name="Arial", size=10, bold=True)

    _auto_width(ws1)

    # ── Sheet 2: Top Productos ──
    ws2 = wb.create_sheet("Top Productos")
    if top20:
        cols2 = list(top20[0].keys())
        _style_header(ws2, cols2)
        for r_idx, row in enumerate(top20, 2):
            for c_idx, col in enumerate(cols2, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=row.get(col))
                _style_cell(cell)
                # Color growth column
                if col == "crecimiento_pct" and row.get(col) is not None:
                    try:
                        val = float(row[col])
                        cell.fill = green_fill if val > 0 else red_fill if val < 0 else PatternFill()
                    except (ValueError, TypeError):
                        pass
        _auto_width(ws2)

    # ── Sheet 3: Competencia ──
    ws3 = wb.create_sheet("Competencia")
    comp_cols = ["Nuestro Producto", "Competidor", "Producto Competidor", "Precio", "Horas", "ECTS", "Titulacion", "Atributos de Valor", "Diferenciador Clave", "URL"]
    _style_header(ws3, comp_cols)
    row_idx = 2
    for res in research_results:
        for comp in res.get("competitors", []):
            vals = [
                res.get("our_product", ""),
                comp.get("competitor_name", ""),
                comp.get("product_name", ""),
                comp.get("price", "N/D"),
                comp.get("hours", "N/D"),
                comp.get("ects", "N/D"),
                comp.get("degree_type", "N/D"),
                comp.get("value_attributes", ""),
                comp.get("key_differentiator", ""),
                comp.get("url", ""),
            ]
            for c_idx, v in enumerate(vals, 1):
                cell = ws3.cell(row=row_idx, column=c_idx, value=v)
                _style_cell(cell)
                if c_idx == 10 and v and v.startswith("http"):
                    cell.hyperlink = v
                    cell.font = Font(name="Arial", size=10, color="2E86AB", underline="single")
            row_idx += 1
    _auto_width(ws3)

    # ── Sheet 4: DAFO ──
    ws4 = wb.create_sheet("DAFO")
    sections = [
        ("Fortalezas", swot.get("strengths", []), green_fill),
        ("Debilidades", swot.get("weaknesses", []), orange_fill),
        ("Oportunidades", swot.get("opportunities", []), blue_fill),
        ("Amenazas", swot.get("threats", []), red_fill),
    ]
    current_row = 1
    for title, items, fill in sections:
        cell = ws4.cell(row=current_row, column=1, value=title)
        cell.font = Font(name="Arial", size=12, bold=True, color="1B2A4A")
        cell.fill = fill
        ws4.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        current_row += 1
        for item in items:
            cell = ws4.cell(row=current_row, column=1, value=f"  - {item}")
            _style_cell(cell)
            ws4.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            current_row += 1
        current_row += 1
    _auto_width(ws4)

    # ── Sheet 5: Propuestas Mejora ──
    ws5 = wb.create_sheet("Propuestas Mejora")
    imp_cols = ["Prioridad", "Producto", "Precio Actual -> Propuesto", "Nombre Propuesto", "Horas/ECTS", "Atributos a Anadir", "Justificacion"]
    _style_header(ws5, imp_cols)
    for i, imp in enumerate(improvements, 2):
        prio_map = {"alta": "Alta", "media": "Media", "baja": "Baja"}
        vals = [
            prio_map.get(imp.get("priority", ""), imp.get("priority", "")),
            imp.get("current_product", ""),
            f"{imp.get('current_price', '')} -> {imp.get('proposed_price', '')}",
            imp.get("proposed_name", "-"),
            imp.get("proposed_hours_ects", "-"),
            ", ".join(imp.get("attributes_to_add", [])),
            imp.get("strategic_justification", ""),
        ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws5.cell(row=i, column=c_idx, value=v)
            _style_cell(cell)
            if c_idx == 1:
                prio = imp.get("priority", "")
                cell.fill = {"alta": red_fill, "media": orange_fill, "baja": green_fill}.get(prio, PatternFill())
    _auto_width(ws5)

    # ── Sheet 6: Nuevos Productos ──
    ws6 = wb.create_sheet("Nuevos Productos")
    np_cols = ["#", "Producto", "Tipo", "Precio", "Horas/ECTS", "Atributos Clave", "Justificacion"]
    _style_header(ws6, np_cols)
    for i, np_ in enumerate(new_products, 2):
        vals = [
            i - 1,
            np_.get("name", ""),
            np_.get("type", ""),
            np_.get("recommended_price", ""),
            np_.get("hours_ects", ""),
            ", ".join(np_.get("key_attributes", [])),
            np_.get("strategic_justification", ""),
        ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws6.cell(row=i, column=c_idx, value=v)
            _style_cell(cell)
    _auto_width(ws6)

    # ── Sheet 7: Estrellas Competencia ──
    ws7 = wb.create_sheet("Estrellas Competencia")
    star_cols = ["Clasificacion", "Competidor", "Producto", "Razon", "Impacto"]
    _style_header(ws7, star_cols)
    cls_map = {"amenaza_directa": "Amenaza directa", "oportunidad_nicho": "Oportunidad", "referente_calidad": "Referente"}
    cls_fill = {"amenaza_directa": red_fill, "oportunidad_nicho": orange_fill, "referente_calidad": green_fill}
    for i, s in enumerate(competitor_stars, 2):
        cls = s.get("classification", "")
        vals = [
            cls_map.get(cls, cls),
            s.get("competitor", ""),
            s.get("product", ""),
            s.get("reason", ""),
            s.get("impact", ""),
        ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws7.cell(row=i, column=c_idx, value=v)
            _style_cell(cell)
            if c_idx == 1:
                cell.fill = cls_fill.get(cls, PatternFill())
    _auto_width(ws7)

    # ── Sheet 8: Mapa Precios-Pais ──
    ws8 = wb.create_sheet("Mapa Precios-Pais")
    map_cols = [
        "Idioma/Area", "Producto Actual", "Institucion Educativa", "Pais Objetivo",
        "Nicho/Razon de Ser", "Nivel Educativo", "Precio (EUR)", "ECTS",
        "Matriculas/Ano", "Total Matriculas", "Ticket Medio Real (EUR)",
        "Se Pisa con Otro?", "Diagnostico",
    ]
    _style_header(ws8, map_cols)
    ws8.cell(row=2, column=1, value="(Datos a completar por el analista con el detalle de la categoria seleccionada)")
    _auto_width(ws8)

    # ── Sheet 9: Demanda Estrategica ──
    ws9 = wb.create_sheet("Demanda Estrategica")
    ws9.cell(row=1, column=1, value="SECCION 1: Inventario por Categoria").font = Font(name="Arial", size=12, bold=True, color="1B2A4A")
    inv_cols = ["Producto", "Tipo", "Precio", "ECTS", "Matriculas/Ano", "Tendencia", "Estado", "Valoracion Estrategica"]
    for c_idx, col in enumerate(inv_cols, 1):
        cell = ws9.cell(row=2, column=c_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
    ws9.cell(row=3, column=1, value="(Datos a completar)")

    row = 6
    ws9.cell(row=row, column=1, value="SECCION 2: Huecos Criticos").font = Font(name="Arial", size=12, bold=True, color="1B2A4A")
    gap_cols = ["Area/Idioma", "Tiene Producto", "Version Universitaria", "Prep. Examen", "Certificacion", "Hueco Detectado", "Prioridad", "Accion Propuesta"]
    for c_idx, col in enumerate(gap_cols, 1):
        cell = ws9.cell(row=row + 1, column=c_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill

    row = 10
    ws9.cell(row=row, column=1, value="SECCION 3: Propuesta Nueva Linea").font = Font(name="Arial", size=12, bold=True, color="1B2A4A")
    line_cols = ["Nombre", "Area", "Tipo", "Precio", "Horas", "ECTS", "Institucion", "Publico", "Prep. Examen", "Certificacion", "Atributos", "Justificacion"]
    for c_idx, col in enumerate(line_cols, 1):
        cell = ws9.cell(row=row + 1, column=c_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill

    row = 14
    ws9.cell(row=row, column=1, value="SECCION 4: Datos de Contexto de Mercado").font = Font(name="Arial", size=12, bold=True, color="1B2A4A")
    ws9.cell(row=row + 1, column=1, value="(Datos de mercado a completar por el analista)")

    _auto_width(ws9)

    wb.save(output_path)
    return output_path


def _truncate(s: str, max_len: int) -> str:
    s = str(s)
    return s[:max_len] + "..." if len(s) > max_len else s
